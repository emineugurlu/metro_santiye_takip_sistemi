import tkinter as tk
from tkinter import ttk, messagebox, filedialog # filedialog burada önemli, dosya kaydetme penceresi için!
from database import Database # Database sınıfınızı kullanmak için gerekli
from datetime import datetime # Rapor dosyalarına tarih/saat eklemek için gerekli
import os # Eğer raporları kaydederken klasör oluşturma/yol işlemleri yapacaksanız ekleyebilirsiniz, zorunlu değil ama iyi bir pratik olabilir.

class MetroSantiyeApp:
    def __init__(self, root): # Burada SADECE 'root' argümanı olmalı
        self.root = root
        self.root.title("Metro Şantiye Takip Sistemi")
        
        # Veritabanı bağlantısını sınıfın KENDİ içinde oluşturun
        # Data klasörünü kontrol et ve oluştur
        import os # Eğer henüz en üstte import etmediyseniz burada da edebilirsiniz
        if not os.path.exists("data"):
            os.makedirs("data")
        
        from database import Database # Database sınıfını burada import ediyoruz
        self.db = Database("data/santiye.db") # Database nesnesi burada oluşturuluyor.
        self.db.create_tables() # Tabloları da burada oluşturalım.
        
        self.create_widgets()

    def create_widgets(self):
        # Sol menü çerçevesi
        self.menu_frame = tk.Frame(self.root, width=200, bg="#2c3e50")
        self.menu_frame.pack(side="left", fill="y")

        # İçerik çerçevesi
        self.content_frame = tk.Frame(self.root, bg="#ecf0f1")
        self.content_frame.pack(side="right", fill="both", expand=True)

        # Menü düğmeleri
        tk.Button(self.menu_frame, text="Personel", command=self.show_personnel_module,
                  font=("Arial", 12), bg="#34495e", fg="white", bd=0, padx=10, pady=10).pack(fill="x", pady=5)
        tk.Button(self.menu_frame, text="Ekipman", command=self.show_equipment_module,
                  font=("Arial", 12), bg="#34495e", fg="white", bd=0, padx=10, pady=10).pack(fill="x", pady=5)
        tk.Button(self.menu_frame, text="Görevler", command=self.show_tasks_module,
                  font=("Arial", 12), bg="#34495e", fg="white", bd=0, padx=10, pady=10).pack(fill="x", pady=5)
        tk.Button(self.menu_frame, text="Acil Durum", command=self.show_emergency_module,
                  font=("Arial", 12), bg="#34495e", fg="white", bd=0, padx=10, pady=10).pack(fill="x", pady=5)
        tk.Button(self.menu_frame, text="Raporlar", command=self.show_reports_module,
                  font=("Arial", 12), bg="#34495e", fg="white", bd=0, padx=10, pady=10).pack(fill="x", pady=5)
        
        # Başlangıçta personel modülünü göster
        self.show_personnel_module()

    def clear_content_frame(self):
        for widget in self.content_frame.winfo_children():
            widget.destroy()

    # --- Personel Modülü ---
    def show_personnel_module(self):
        self.clear_content_frame()
        tk.Label(self.content_frame, text="Personel Takip Modülü", font=("Arial", 16, "bold"), bg="#ecf0f1").pack(pady=20)

        # Personel listesi için Treeview
        columns = ("ID", "Ad", "Görev", "Vardiya")
        self.personnel_tree = ttk.Treeview(self.content_frame, columns=columns, show="headings")
        for col in columns:
            self.personnel_tree.heading(col, text=col)
            self.personnel_tree.column(col, width=100)
        self.personnel_tree.pack(pady=10, padx=10, fill="both", expand=True)

        self.load_personnel_data()

        # Düğmeler
        button_frame = tk.Frame(self.content_frame, bg="#ecf0f1")
        button_frame.pack(pady=10)
        tk.Button(button_frame, text="Yeni Personel Ekle", command=self.add_personnel).pack(side="left", padx=5)
        tk.Button(button_frame, text="Personel Düzenle", command=self.edit_personnel).pack(side="left", padx=5)
        tk.Button(button_frame, text="Personel Sil", command=self.delete_personnel).pack(side="left", padx=5)

    def load_personnel_data(self):
        for row in self.personnel_tree.get_children():
            self.personnel_tree.delete(row)
        personnel = self.db.fetch_all("SELECT id, name, role, shift FROM personnel")
        for p in personnel:
            self.personnel_tree.insert("", "end", values=p)

    def add_personnel(self):
        # Yeni personel ekleme penceresi/dialogu
        add_window = tk.Toplevel(self.root)
        add_window.title("Yeni Personel Ekle")
        add_window.geometry("300x250") # Pencere boyutunu biraz büyüttüm

        # Etiketler ve Giriş Alanları
        tk.Label(add_window, text="Ad:").pack(pady=5)
        name_entry = tk.Entry(add_window)
        name_entry.pack(pady=5, padx=10, fill="x")

        tk.Label(add_window, text="Görev:").pack(pady=5)
        role_entry = tk.Entry(add_window)
        role_entry.pack(pady=5, padx=10, fill="x")

        tk.Label(add_window, text="Vardiya:").pack(pady=5)
        shift_entry = tk.Entry(add_window)
        shift_entry.pack(pady=5, padx=10, fill="x")

        # Kaydet butonu
        def save():
            name = name_entry.get().strip() # Boşlukları temizle
            role = role_entry.get().strip()
            shift = shift_entry.get().strip()

            if name and role: # Ad ve Görev alanlarının boş olup olmadığını kontrol et
                self.db.execute_query("INSERT INTO personnel (name, role, shift) VALUES (?, ?, ?)", (name, role, shift))
                messagebox.showinfo("Başarılı", "Personel başarıyla eklendi.")
                self.load_personnel_data() # Personel listesini yenile
                add_window.destroy() # Pencereyi kapat
            else:
                messagebox.showerror("Hata", "Ad ve Görev alanları boş bırakılamaz.")

        tk.Button(add_window, text="Kaydet", command=save, bg="#28a745", fg="white", font=("Arial", 10, "bold")).pack(pady=10)
        tk.Button(add_window, text="İptal", command=add_window.destroy, bg="#dc3545", fg="white", font=("Arial", 10)).pack(pady=5)

    def edit_personnel(self):
        selected_item = self.personnel_tree.focus() # Seçili öğeyi al
        if not selected_item: # Hiçbir öğe seçili değilse uyar
            messagebox.showwarning("Seçim Yok", "Lütfen düzenlemek istediğiniz bir personel seçin.")
            return

        values = self.personnel_tree.item(selected_item, 'values') # Seçili öğenin değerlerini al
        person_id = values[0] # ID ilk değer

        # Düzenleme penceresi oluştur
        edit_window = tk.Toplevel(self.root)
        edit_window.title("Personel Düzenle")
        edit_window.geometry("300x250") # Pencere boyutunu ayarla

        # Etiketler ve Giriş Alanları, mevcut değerlerle doldur
        tk.Label(edit_window, text="Ad:").pack(pady=5)
        name_entry = tk.Entry(edit_window)
        name_entry.insert(0, values[1]) # Mevcut adı önceden doldur
        name_entry.pack(pady=5, padx=10, fill="x")

        tk.Label(edit_window, text="Görev:").pack(pady=5)
        role_entry = tk.Entry(edit_window)
        role_entry.insert(0, values[2]) # Mevcut görevi önceden doldur
        role_entry.pack(pady=5, padx=10, fill="x")

        tk.Label(edit_window, text="Vardiya:").pack(pady=5)
        shift_entry = tk.Entry(edit_window)
        shift_entry.insert(0, values[3]) # Mevcut vardiyayı önceden doldur
        shift_entry.pack(pady=5, padx=10, fill="x")

        # Kaydet butonu
        def save():
            name = name_entry.get().strip()
            role = role_entry.get().strip()
            shift = shift_entry.get().strip()

            if name and role: # Ad ve Görev alanlarının boş olup olmadığını kontrol et
                self.db.execute_query("UPDATE personnel SET name=?, role=?, shift=? WHERE id=?", (name, role, shift, person_id))
                messagebox.showinfo("Başarılı", "Personel başarıyla güncellendi.")
                self.load_personnel_data() # Personel listesini yenile
                edit_window.destroy() # Pencereyi kapat
            else:
                messagebox.showerror("Hata", "Ad ve Görev alanları boş bırakılamaz.")

        tk.Button(edit_window, text="Kaydet", command=save, bg="#28a745", fg="white", font=("Arial", 10, "bold")).pack(pady=10)
        tk.Button(edit_window, text="İptal", command=edit_window.destroy, bg="#dc3545", fg="white", font=("Arial", 10)).pack(pady=5)

    def delete_personnel(self):
        selected_item = self.personnel_tree.focus() # Seçili öğeyi al
        if not selected_item: # Hiçbir öğe seçili değilse uyar
            messagebox.showwarning("Seçim Yok", "Lütfen silmek istediğiniz bir personel seçin.")
            return

        values = self.personnel_tree.item(selected_item, 'values') # Seçili öğenin değerlerini al
        person_id = values[0]
        person_name = values[1]

        # Kullanıcıya onay sor
        if messagebox.askyesno("Onay", f"'{person_name}' ({person_id} ID'li) personeli silmek istediğinizden emin misiniz? Bu işlem geri alınamaz."):
            # Personeli sil
            # FOREIGN KEY ON DELETE SET NULL yaptığımız için ekipman/görev güncellemeleri otomatik olacak.
            self.db.execute_query("DELETE FROM personnel WHERE id=?", (person_id,))
            messagebox.showinfo("Başarılı", "Personel başarıyla silindi.")
            self.load_personnel_data() # Personel listesini yenile
        else:
            messagebox.showinfo("İptal Edildi", "Personel silme işlemi iptal edildi.")

    # --- Ekipman Modülü ---
    def show_equipment_module(self):
        self.clear_content_frame()
        tk.Label(self.content_frame, text="Ekipman Zimmet Modülü", font=("Arial", 16, "bold"), bg="#ecf0f1").pack(pady=20)

        # Ekipman listesi için Treeview
        columns = ("ID", "Ad/Tür", "Zimmetli Kişi (ID)", "Durum")
        self.equipment_tree = ttk.Treeview(self.content_frame, columns=columns, show="headings")
        for col in columns:
            self.equipment_tree.heading(col, text=col)
            self.equipment_tree.column(col, width=120)
        self.equipment_tree.pack(pady=10, padx=10, fill="both", expand=True)

        self.load_equipment_data() # Verileri yükle

        # Düğmeler
        button_frame = tk.Frame(self.content_frame, bg="#ecf0f1")
        button_frame.pack(pady=10)
        tk.Button(button_frame, text="Yeni Ekipman Ekle", command=self.add_equipment).pack(side="left", padx=5)
        tk.Button(button_frame, text="Ekipman Düzenle", command=self.edit_equipment).pack(side="left", padx=5)
        tk.Button(button_frame, text="Ekipman Sil", command=self.delete_equipment).pack(side="left", padx=5)
        tk.Button(button_frame, text="Ekipman Zimmetle", command=self.assign_equipment).pack(side="left", padx=10) # Daha geniş ayrım
        tk.Button(button_frame, text="Zimmeti Kaldır", command=self.unassign_equipment).pack(side="left", padx=5)

    def load_equipment_data(self):
        for row in self.equipment_tree.get_children():
            self.equipment_tree.delete(row)
        
        # Ekipman ve zimmetli olduğu personelin adı (JOIN ile)
        equipment = self.db.fetch_all("""
            SELECT 
                e.id, 
                e.name, 
                COALESCE(p.name || ' (ID: ' || p.id || ')', 'Yok') AS assigned_person_info, -- Eğer assigned_to NULL ise 'Yok' göster
                e.status 
            FROM equipment e
            LEFT JOIN personnel p ON e.assigned_to = p.id
        """)
        for eq in equipment:
            self.equipment_tree.insert("", "end", values=eq)

    # --- Ekipman Ekleme ---
    def add_equipment(self):
        add_window = tk.Toplevel(self.root)
        add_window.title("Yeni Ekipman Ekle")
        add_window.geometry("300x200")

        tk.Label(add_window, text="Ekipman Adı/Türü:").pack(pady=5)
        name_entry = tk.Entry(add_window)
        name_entry.pack(pady=5, padx=10, fill="x")

        tk.Label(add_window, text="Durum:").pack(pady=5)
        # Durum için seçenekli bir Combobox kullanabiliriz
        status_options = ["Mevcut", "Zimmetli", "Arızalı", "Bakımda"]
        status_combobox = ttk.Combobox(add_window, values=status_options)
        status_combobox.set("Mevcut") # Varsayılan değer
        status_combobox.pack(pady=5, padx=10, fill="x")

        def save():
            name = name_entry.get().strip()
            status = status_combobox.get().strip()
            if name and status:
                self.db.execute_query("INSERT INTO equipment (name, status) VALUES (?, ?)", (name, status))
                messagebox.showinfo("Başarılı", "Ekipman başarıyla eklendi.")
                self.load_equipment_data()
                add_window.destroy()
            else:
                messagebox.showerror("Hata", "Ekipman Adı/Türü ve Durum alanları boş bırakılamaz.")

        tk.Button(add_window, text="Kaydet", command=save, bg="#28a745", fg="white", font=("Arial", 10, "bold")).pack(pady=10)
        tk.Button(add_window, text="İptal", command=add_window.destroy, bg="#dc3545", fg="white", font=("Arial", 10)).pack(pady=5)

    # --- Ekipman Düzenleme ---
    def edit_equipment(self):
        selected_item = self.equipment_tree.focus()
        if not selected_item:
            messagebox.showwarning("Seçim Yok", "Lütfen düzenlemek istediğiniz bir ekipman seçin.")
            return

        values = self.equipment_tree.item(selected_item, 'values')
        equipment_id = values[0]
        
        edit_window = tk.Toplevel(self.root)
        edit_window.title("Ekipman Düzenle")
        edit_window.geometry("300x200")

        tk.Label(edit_window, text="Ekipman Adı/Türü:").pack(pady=5)
        name_entry = tk.Entry(edit_window)
        name_entry.insert(0, values[1])
        name_entry.pack(pady=5, padx=10, fill="x")

        tk.Label(edit_window, text="Durum:").pack(pady=5)
        status_options = ["Mevcut", "Zimmetli", "Arızalı", "Bakımda"]
        status_combobox = ttk.Combobox(edit_window, values=status_options)
        status_combobox.set(values[3]) # Mevcut durumu seç
        status_combobox.pack(pady=5, padx=10, fill="x")

        def save():
            name = name_entry.get().strip()
            status = status_combobox.get().strip()
            if name and status:
                self.db.execute_query("UPDATE equipment SET name=?, status=? WHERE id=?", (name, status, equipment_id))
                messagebox.showinfo("Başarılı", "Ekipman başarıyla güncellendi.")
                self.load_equipment_data()
                edit_window.destroy()
            else:
                messagebox.showerror("Hata", "Ekipman Adı/Türü ve Durum alanları boş bırakılamaz.")

        tk.Button(edit_window, text="Kaydet", command=save, bg="#28a745", fg="white", font=("Arial", 10, "bold")).pack(pady=10)
        tk.Button(edit_window, text="İptal", command=edit_window.destroy, bg="#dc3545", fg="white", font=("Arial", 10)).pack(pady=5)

    # --- Ekipman Silme ---
    def delete_equipment(self):
        selected_item = self.equipment_tree.focus()
        if not selected_item:
            messagebox.showwarning("Seçim Yok", "Lütfen silmek istediğiniz bir ekipman seçin.")
            return

        values = self.equipment_tree.item(selected_item, 'values')
        equipment_id = values[0]
        equipment_name = values[1]

        if messagebox.askyesno("Onay", f"'{equipment_name}' ({equipment_id} ID'li) ekipmanı silmek istediğinizden emin misiniz? Bu işlem geri alınamaz."):
            self.db.execute_query("DELETE FROM equipment WHERE id=?", (equipment_id,))
            messagebox.showinfo("Başarılı", "Ekipman başarıyla silindi.")
            self.load_equipment_data()
        else:
            messagebox.showinfo("İptal Edildi", "Ekipman silme işlemi iptal edildi.")

    # --- Ekipman Zimmetleme ---
    def assign_equipment(self):
        selected_item = self.equipment_tree.focus()
        if not selected_item:
            messagebox.showwarning("Seçim Yok", "Lütfen zimmetlemek istediğiniz bir ekipman seçin.")
            return

        values = self.equipment_tree.item(selected_item, 'values')
        equipment_id = values[0]
        current_status = values[3]

        if current_status == "Zimmetli":
            messagebox.showwarning("Uyarı", "Bu ekipman zaten zimmetli.")
            return

        # Personel seçme penceresi
        assign_window = tk.Toplevel(self.root)
        assign_window.title("Ekipmanı Zimmetle")
        assign_window.geometry("400x300")

        tk.Label(assign_window, text=f"'{values[1]}' adlı ekipmanı kime zimmetleyeceksiniz?", font=("Arial", 10, "bold")).pack(pady=10)

        # Personel listesi için Treeview
        columns = ("ID", "Ad", "Görev")
        personnel_select_tree = ttk.Treeview(assign_window, columns=columns, show="headings")
        for col in columns:
            personnel_select_tree.heading(col, text=col)
            personnel_select_tree.column(col, width=100)
        
        # Sadece tekli seçim yapılmasını sağla (varsayılan genelde "browse" modudur)
        personnel_select_tree.config(selectmode="browse") 

        personnel_select_tree.pack(pady=10, padx=10, fill="both", expand=True)

        # Personel verilerini yükle
        personnel_data = self.db.fetch_all("SELECT id, name, role FROM personnel")
        for p in personnel_data:
            personnel_select_tree.insert("", "end", values=p)

        def save_assignment():
            # Burası önemli: get_selection() veya focus() kullanabiliriz
            # focus() tekli seçimde seçili öğeyi döndürür.
            selected_person_item = personnel_select_tree.focus() 
            
            if not selected_person_item:
                messagebox.showerror("Hata", "Lütfen ekipmanı zimmetleyeceğiniz bir personel seçin.")
                return
            
            # Seçili öğenin verilerini al
            person_values = personnel_select_tree.item(selected_person_item, 'values')
            
            # person_values'ın boş veya eksik olup olmadığını kontrol edelim
            if not person_values or len(person_values) < 2: # En az ID ve Ad olmalı
                messagebox.showerror("Hata", "Seçilen personel verileri alınamadı. Lütfen tekrar deneyin.")
                return

            try:
                person_id = int(person_values[0]) # ID'nin int olduğundan emin olalım
                person_name = person_values[1]

                if messagebox.askyesno("Onay", f"'{values[1]}' ekipmanını '{person_name}' ({person_id} ID'li) kişiye zimmetlemek istediğinizden emin misiniz?"):
                    self.db.execute_query("UPDATE equipment SET assigned_to=?, status='Zimmetli' WHERE id=?", (person_id, equipment_id))
                    messagebox.showinfo("Başarılı", f"Ekipman '{person_name}' kişiye başarıyla zimmetlendi.")
                    self.load_equipment_data()
                    assign_window.destroy()
                else:
                    messagebox.showinfo("İptal Edildi", "Zimmetleme işlemi iptal edildi.")
            except ValueError:
                messagebox.showerror("Hata", "Personel ID'si geçerli değil.")
            except Exception as e:
                messagebox.showerror("Hata", f"Zimmetleme sırasında bir hata oluştu: {e}")


        tk.Button(assign_window, text="Zimmetle", command=save_assignment, bg="#28a745", fg="white", font=("Arial", 10, "bold")).pack(pady=10)
        tk.Button(assign_window, text="İptal", command=assign_window.destroy, bg="#dc3545", fg="white", font=("Arial", 10)).pack(pady=5)

    # --- Zimmeti Kaldırma ---
    def unassign_equipment(self):
        selected_item = self.equipment_tree.focus()
        if not selected_item:
            messagebox.showwarning("Seçim Yok", "Lütfen zimmeti kaldırılacak bir ekipman seçin.")
            return

        values = self.equipment_tree.item(selected_item, 'values')
        equipment_id = values[0]
        assigned_person_info = values[2] # "Ad (ID: X)" veya "Yok"

        if assigned_person_info == "Yok":
            messagebox.showwarning("Uyarı", "Bu ekipman zaten zimmetli değil.")
            return

        if messagebox.askyesno("Onay", f"'{values[1]}' ekipmanının zimmetini kaldırmak istediğinizden emin misiniz?"):
            self.db.execute_query("UPDATE equipment SET assigned_to=NULL, status='Mevcut' WHERE id=?", (equipment_id,))
            messagebox.showinfo("Başarılı", "Ekipman zimmeti başarıyla kaldırıldı.")
            self.load_equipment_data()
        else:
            messagebox.showinfo("İptal Edildi", "Zimmeti kaldırma işlemi iptal edildi.")

    # --- Görevler Modülü ---
    def show_tasks_module(self):
        self.clear_content_frame()
        tk.Label(self.content_frame, text="Görev Yönetim Modülü", font=("Arial", 16, "bold"), bg="#ecf0f1").pack(pady=20)

        # Görev listesi için Treeview
        columns = ("ID", "Görev Adı", "Atanan Kişi (ID)", "Bitiş Tarihi", "Durum")
        self.tasks_tree = ttk.Treeview(self.content_frame, columns=columns, show="headings")
        for col in columns:
            self.tasks_tree.heading(col, text=col)
            self.tasks_tree.column(col, width=120)
        self.tasks_tree.pack(pady=10, padx=10, fill="both", expand=True)

        self.load_tasks_data() # Verileri yükle

        # Düğmeler
        button_frame = tk.Frame(self.content_frame, bg="#ecf0f1")
        button_frame.pack(pady=10)
        tk.Button(button_frame, text="Yeni Görev Ekle", command=self.add_task).pack(side="left", padx=5)
        tk.Button(button_frame, text="Görevi Düzenle", command=self.edit_task).pack(side="left", padx=5)
        tk.Button(button_frame, text="Görevi Sil", command=self.delete_task).pack(side="left", padx=5)
        tk.Button(button_frame, text="Tamamlandı Olarak İşaretle", command=self.mark_task_completed).pack(side="left", padx=10)
        
    def load_tasks_data(self):
        for row in self.tasks_tree.get_children():
            self.tasks_tree.delete(row)
        
        # Görev ve atandığı personelin adı (JOIN ile)
        tasks = self.db.fetch_all("""
            SELECT 
                t.id, 
                t.task_name, 
                COALESCE(p.name || ' (ID: ' || p.id || ')', 'Atanmadı') AS assigned_person_info, -- Eğer assigned_person NULL ise 'Atanmadı' göster
                t.deadline,
                t.status
            FROM tasks t
            LEFT JOIN personnel p ON t.assigned_person = p.id
        """)
        for task in tasks:
            # Duruma göre renkler atayabiliriz (isteğe bağlı, sonra da yapılabilir)
            # Örneğin, 'Tamamlandı' için yeşil, 'Beklemede' için turuncu vb.
            # Şu an için sadece ekleyelim:
            self.tasks_tree.insert("", "end", values=task)

    # --- Görev Ekleme ---
    def add_task(self):
        add_window = tk.Toplevel(self.root)
        add_window.title("Yeni Görev Ekle")
        add_window.geometry("400x350")

        tk.Label(add_window, text="Görev Adı:").pack(pady=5)
        task_name_entry = tk.Entry(add_window)
        task_name_entry.pack(pady=5, padx=10, fill="x")

        tk.Label(add_window, text="Bitiş Tarihi (YYYY-MM-DD):").pack(pady=5)
        deadline_entry = tk.Entry(add_window)
        deadline_entry.pack(pady=5, padx=10, fill="x")
        # Örnek tarih formatı için placeholder
        deadline_entry.insert(0, "YYYY-MM-DD")
        deadline_entry.bind("<FocusIn>", lambda event: deadline_entry.delete(0, "end") if deadline_entry.get() == "YYYY-MM-DD" else None)

        tk.Label(add_window, text="Durum:").pack(pady=5)
        status_options = ["Beklemede", "Devam Ediyor", "Tamamlandı", "İptal Edildi"]
        status_combobox = ttk.Combobox(add_window, values=status_options)
        status_combobox.set("Beklemede") # Varsayılan değer
        status_combobox.pack(pady=5, padx=10, fill="x")

        tk.Label(add_window, text="Atanacak Personel:").pack(pady=5)
        
        # Personel seçimi için Combobox
        # Önce tüm personeli çekelim
        personnel_data = self.db.fetch_all("SELECT id, name FROM personnel")
        personnel_dict = {f"{p[1]} (ID: {p[0]})": p[0] for p in personnel_data} # "Ad (ID: X)": ID
        personnel_names = list(personnel_dict.keys())
        personnel_names.insert(0, "Atama Yok") # İlk seçenek olarak atanmamış olsun

        assigned_person_combobox = ttk.Combobox(add_window, values=personnel_names)
        assigned_person_combobox.set("Atama Yok")
        assigned_person_combobox.pack(pady=5, padx=10, fill="x")

        def save():
            task_name = task_name_entry.get().strip()
            deadline = deadline_entry.get().strip()
            status = status_combobox.get().strip()
            selected_person_text = assigned_person_combobox.get()
            
            assigned_person_id = None
            if selected_person_text != "Atama Yok":
                try:
                    # 'Ad (ID: X)' formatından ID'yi çek
                    person_id_str = selected_person_text.split('(ID: ')[1].rstrip(')')
                    assigned_person_id = int(person_id_str)
                except (IndexError, ValueError):
                    messagebox.showerror("Hata", "Geçersiz personel seçimi formatı.")
                    return

            if task_name and deadline and status:
                self.db.execute_query(
                    "INSERT INTO tasks (task_name, assigned_person, deadline, status) VALUES (?, ?, ?, ?)", 
                    (task_name, assigned_person_id, deadline, status)
                )
                messagebox.showinfo("Başarılı", "Görev başarıyla eklendi.")
                self.load_tasks_data()
                add_window.destroy()
            else:
                messagebox.showerror("Hata", "Görev Adı, Bitiş Tarihi ve Durum alanları boş bırakılamaz.")

        tk.Button(add_window, text="Kaydet", command=save, bg="#28a745", fg="white", font=("Arial", 10, "bold")).pack(pady=10)
        tk.Button(add_window, text="İptal", command=add_window.destroy, bg="#dc3545", fg="white", font=("Arial", 10)).pack(pady=5)

    # --- Görev Düzenleme ---
    def edit_task(self):
        selected_item = self.tasks_tree.focus()
        if not selected_item:
            messagebox.showwarning("Seçim Yok", "Lütfen düzenlemek istediğiniz bir görev seçin.")
            return

        values = self.tasks_tree.item(selected_item, 'values')
        task_id = values[0]
        
        edit_window = tk.Toplevel(self.root)
        edit_window.title("Görevi Düzenle")
        edit_window.geometry("400x350")

        tk.Label(edit_window, text="Görev Adı:").pack(pady=5)
        task_name_entry = tk.Entry(edit_window)
        task_name_entry.insert(0, values[1])
        task_name_entry.pack(pady=5, padx=10, fill="x")

        tk.Label(edit_window, text="Bitiş Tarihi (YYYY-MM-DD):").pack(pady=5)
        deadline_entry = tk.Entry(edit_window)
        deadline_entry.insert(0, values[3])
        deadline_entry.pack(pady=5, padx=10, fill="x")

        tk.Label(edit_window, text="Durum:").pack(pady=5)
        status_options = ["Beklemede", "Devam Ediyor", "Tamamlandı", "İptal Edildi"]
        status_combobox = ttk.Combobox(edit_window, values=status_options)
        status_combobox.set(values[4]) # Mevcut durumu seç
        status_combobox.pack(pady=5, padx=10, fill="x")

        tk.Label(edit_window, text="Atanacak Personel:").pack(pady=5)
        personnel_data = self.db.fetch_all("SELECT id, name FROM personnel")
        personnel_dict = {f"{p[1]} (ID: {p[0]})": p[0] for p in personnel_data}
        personnel_names = list(personnel_dict.keys())
        personnel_names.insert(0, "Atama Yok")

        assigned_person_combobox = ttk.Combobox(edit_window, values=personnel_names)
        # Mevcut atanan kişiyi combobox'ta seçili hale getir
        current_assigned = values[2] # 'Ad (ID: X)' veya 'Atanmadı'
        if current_assigned == "Atanmadı":
            assigned_person_combobox.set("Atama Yok")
        else:
            # Sadece 'Ad (ID: X)' kısmını al ve combobox'ta ara
            # Eğer tam eşleşme yoksa varsayılana düşsün
            if current_assigned in personnel_names:
                assigned_person_combobox.set(current_assigned)
            else:
                assigned_person_combobox.set("Atama Yok") # Bulunamazsa varsayılan

        assigned_person_combobox.pack(pady=5, padx=10, fill="x")

        def save():
            task_name = task_name_entry.get().strip()
            deadline = deadline_entry.get().strip()
            status = status_combobox.get().strip()
            selected_person_text = assigned_person_combobox.get()
            
            assigned_person_id = None
            if selected_person_text != "Atama Yok":
                try:
                    person_id_str = selected_person_text.split('(ID: ')[1].rstrip(')')
                    assigned_person_id = int(person_id_str)
                except (IndexError, ValueError):
                    messagebox.showerror("Hata", "Geçersiz personel seçimi formatı.")
                    return

            if task_name and deadline and status:
                self.db.execute_query(
                    "UPDATE tasks SET task_name=?, assigned_person=?, deadline=?, status=? WHERE id=?", 
                    (task_name, assigned_person_id, deadline, status, task_id)
                )
                messagebox.showinfo("Başarılı", "Görev başarıyla güncellendi.")
                self.load_tasks_data()
                edit_window.destroy()
            else:
                messagebox.showerror("Hata", "Görev Adı, Bitiş Tarihi ve Durum alanları boş bırakılamaz.")

        tk.Button(edit_window, text="Kaydet", command=save, bg="#28a745", fg="white", font=("Arial", 10, "bold")).pack(pady=10)
        tk.Button(edit_window, text="İptal", command=edit_window.destroy, bg="#dc3545", fg="white", font=("Arial", 10)).pack(pady=5)

    # --- Görev Silme ---
    def delete_task(self):
        selected_item = self.tasks_tree.focus()
        if not selected_item:
            messagebox.showwarning("Seçim Yok", "Lütfen silmek istediğiniz bir görev seçin.")
            return

        values = self.tasks_tree.item(selected_item, 'values')
        task_id = values[0]
        task_name = values[1]

        if messagebox.askyesno("Onay", f"'{task_name}' ({task_id} ID'li) görevi silmek istediğinizden emin misiniz? Bu işlem geri alınamaz."):
            self.db.execute_query("DELETE FROM tasks WHERE id=?", (task_id,))
            messagebox.showinfo("Başarılı", "Görev başarıyla silindi.")
            self.load_tasks_data()
        else:
            messagebox.showinfo("İptal Edildi", "Görev silme işlemi iptal edildi.")

    # --- Görevi Tamamlandı Olarak İşaretle ---
    def mark_task_completed(self):
        selected_item = self.tasks_tree.focus()
        if not selected_item:
            messagebox.showwarning("Seçim Yok", "Lütfen tamamlandı olarak işaretlemek istediğiniz bir görev seçin.")
            return

        values = self.tasks_tree.item(selected_item, 'values')
        task_id = values[0]
        current_status = values[4]

        if current_status == "Tamamlandı":
            messagebox.showwarning("Uyarı", "Bu görev zaten tamamlandı olarak işaretlenmiş.")
            return

        if messagebox.askyesno("Onay", f"'{values[1]}' görevini tamamlandı olarak işaretlemek istediğinizden emin misiniz?"):
            self.db.execute_query("UPDATE tasks SET status='Tamamlandı' WHERE id=?", (task_id,))
            messagebox.showinfo("Başarılı", "Görev başarıyla tamamlandı olarak işaretlendi.")
            self.load_tasks_data()
        else:
            messagebox.showinfo("İptal Edildi", "Görev durumu güncelleme işlemi iptal edildi.")

   # --- Acil Durum Bildirim Modülü ---
    def show_emergency_module(self):
        self.clear_content_frame()
        # Ana içerik çerçevesi arka plan rengini ayarla
        self.content_frame.config(bg="#ecf0f1") # Arka plan rengini ekledim
        
        # Başlık etiketi
        tk.Label(self.content_frame, text="Acil Durum Bildirim Modülü", font=("Arial", 16, "bold"), bg="#ecf0f1", fg="#2c3e50").pack(pady=20) # Renk ve font ekledim

        # Acil durum listesi için Treeview
        columns = ("ID", "Tarih", "Saat", "Tür", "Açıklama", "Etkilenen Personel", "Etkilenen Ekipman", "Alınan Aksiyon")
        self.emergency_tree = ttk.Treeview(self.content_frame, columns=columns, show="headings")
        
        # Sütun başlıkları ve genişlikleri
        self.emergency_tree.heading("ID", text="ID")
        self.emergency_tree.column("ID", width=40, anchor="center")
        self.emergency_tree.heading("Tarih", text="Tarih")
        self.emergency_tree.column("Tarih", width=100, anchor="center")
        self.emergency_tree.heading("Saat", text="Saat")
        self.emergency_tree.column("Saat", width=70, anchor="center")
        self.emergency_tree.heading("Tür", text="Tür")
        self.emergency_tree.column("Tür", width=120)
        self.emergency_tree.heading("Açıklama", text="Açıklama")
        self.emergency_tree.column("Açıklama", width=200)
        self.emergency_tree.heading("Etkilenen Personel", text="Etkilenen Personel")
        self.emergency_tree.column("Etkilenen Personel", width=180)
        self.emergency_tree.heading("Etkilenen Ekipman", text="Etkilenen Ekipman")
        self.emergency_tree.column("Etkilenen Ekipman", width=180)
        self.emergency_tree.heading("Alınan Aksiyon", text="Alınan Aksiyon")
        self.emergency_tree.column("Alınan Aksiyon", width=150)

        self.emergency_tree.pack(pady=10, padx=10, fill="both", expand=True)

        self.load_emergency_data() # Verileri yükle

        # Düğmeler
        button_frame = tk.Frame(self.content_frame, bg="#ecf0f1") # Arka plan rengini eşitledim
        button_frame.pack(pady=10)
        
        # Buton stillerini verilen tasarıma uygun hale getirdim
        tk.Button(button_frame, text="Yeni Acil Durum Kaydı", command=self.add_emergency,
                  bg="#34495e", fg="white", font=("Arial", 10, "bold"), padx=10, pady=5, bd=0, relief="flat").pack(side="left", padx=5)
        tk.Button(button_frame, text="Acil Durum Düzenle", command=self.edit_emergency,
                  bg="#34495e", fg="white", font=("Arial", 10, "bold"), padx=10, pady=5, bd=0, relief="flat").pack(side="left", padx=5)
        tk.Button(button_frame, text="Acil Durum Sil", command=self.delete_emergency,
                  bg="#e74c3c", fg="white", font=("Arial", 10, "bold"), padx=10, pady=5, bd=0, relief="flat").pack(side="left", padx=5)
    
    def load_emergency_data(self):
        # Treeview'i temizle
        for i in self.emergency_tree.get_children():
            self.emergency_tree.delete(i)

        # Veritabanından acil durum verilerini çek
        # Personel ve ekipman adlarını da çekmek için JOIN kullanıyoruz
        emergency_data = self.db.fetch_all("""
            SELECT
                ee.id,
                ee.date,
                ee.time,
                ee.event_type,
                ee.description,
                -- Personel bilgisi için CASE WHEN kontrolü: ID varsa Ad Soyad (ID: X) formatında, yoksa boş string
                CASE
                    WHEN p.id IS NOT NULL THEN p.name || ' ' || p.surname || ' (ID: ' || p.id || ')'
                    ELSE ''
                END AS affected_personnel_display,
                -- Ekipman bilgisi için CASE WHEN kontrolü: ID varsa Ekipman Adı (ID: X) formatında, yoksa boş string
                CASE
                    WHEN eq.id IS NOT NULL THEN eq.name || ' (ID: ' || eq.id || ')'
                    ELSE ''
                END AS affected_equipment_display,
                ee.action_taken
            FROM
                emergency_events ee
            LEFT JOIN
                personnel p ON ee.affected_personnel_ids = p.id
            LEFT JOIN
                equipment eq ON ee.affected_equipment_ids = eq.id
            ORDER BY
                ee.date DESC, ee.time DESC
        """)

        # Verileri Treeview'e ekle
        for row in emergency_data:
            self.emergency_tree.insert("", "end", values=row)

    def add_emergency(self):
        # Yeni Acil Durum Kaydı için pencere oluştur
        add_window = tk.Toplevel(self.root)
        add_window.title("Yeni Acil Durum Kaydı Ekle")
        add_window.geometry("400x550") # Tasarıma uygun boyut
        add_window.transient(self.root) # Ana pencerenin üzerinde kalmasını sağlar
        add_window.grab_set() # Ana pencereyle etkileşimi engeller
        add_window.configure(bg="#ecf0f1") # Arka plan rengini ekledim

        # Etiket ve giriş alanı stilleri için varsayılan font ve genişlikleri kullanıyorum.
        # Label'lar için arka plan rengi eklendi.
        label_font = ("Arial", 10)
        entry_font = ("Arial", 10)
        entry_width = 37 # Tasarıma uygun genişlik
        pady_value = 5 # Boşluk değeri

        tk.Label(add_window, text="Tarih (YYYY-MM-DD):", font=label_font, bg="#ecf0f1").pack(pady=(10,0))
        date_entry = tk.Entry(add_window, font=entry_font, width=entry_width)
        date_entry.insert(0, datetime.now().strftime("%Y-%m-%d")) # Otomatik bugünün tarihini ekle
        date_entry.pack(pady=pady_value)

        tk.Label(add_window, text="Saat (HH:MM):", font=label_font, bg="#ecf0f1").pack(pady=(10,0))
        time_entry = tk.Entry(add_window, font=entry_font, width=entry_width)
        time_entry.insert(0, datetime.now().strftime("%H:%M")) # Otomatik şu anki saati ekle
        time_entry.pack(pady=pady_value)

        tk.Label(add_window, text="Acil Durum Türü:", font=label_font, bg="#ecf0f1").pack(pady=(10,0))
        event_type_entry = tk.Entry(add_window, font=entry_font, width=entry_width)
        event_type_entry.pack(pady=pady_value)

        tk.Label(add_window, text="Açıklama:", font=label_font, bg="#ecf0f1").pack(pady=(10,0))
        description_entry = tk.Entry(add_window, font=entry_font, width=entry_width)
        description_entry.pack(pady=pady_value)

        # --- Etkilenen Personel Combobox ---
        tk.Label(add_window, text="Etkilenen Personel:", font=label_font, bg="#ecf0f1").pack(pady=(10,0))
        personnel_data = self.db.fetch_all("SELECT id, name,role,shift FROM personnel")
        personnel_options = [""] + [f"{p[1]} {p[2]} (ID: {p[0]})" for p in personnel_data] # Personel adlarını ve ID'lerini çek, başına boş seçenek ekle
        personnel_combobox = ttk.Combobox(add_window, values=personnel_options, state="readonly", width=entry_width - 2) # Combobox genişliği Entry'den biraz küçük olabilir
        personnel_combobox.pack(pady=pady_value)
        personnel_combobox.set("") # Başlangıçta boş seçenek seçili

        # --- Etkilenen Ekipman Combobox ---
        tk.Label(add_window, text="Etkilenen Ekipman:", font=label_font, bg="#ecf0f1").pack(pady=(10,0))
        equipment_data = self.db.fetch_all("SELECT id, name FROM equipment")
        equipment_options = [""] + [f"{e[1]} (ID: {e[0]})" for e in equipment_data] # Ekipman adlarını ve ID'lerini çek, başına boş seçenek ekle
        equipment_combobox = ttk.Combobox(add_window, values=equipment_options, state="readonly", width=entry_width - 2)
        equipment_combobox.pack(pady=pady_value)
        equipment_combobox.set("") # Başlangıçta boş seçenek seçili

        tk.Label(add_window, text="Alınan Aksiyon:", font=label_font, bg="#ecf0f1").pack(pady=(10,0))
        action_taken_entry = tk.Entry(add_window, font=entry_font, width=entry_width)
        action_taken_entry.pack(pady=pady_value)

        def save_emergency():
            date = date_entry.get()
            time = time_entry.get()
            event_type = event_type_entry.get()
            description = description_entry.get()

            # Combobox'tan seçilen değeri al ve ID'yi çıkar
            selected_personnel_str = personnel_combobox.get()
            personnel_id = None
            if selected_personnel_str:
                parts = selected_personnel_str.split("(ID: ")
                if len(parts) > 1:
                    personnel_id = parts[1].strip(")")

            selected_equipment_str = equipment_combobox.get()
            equipment_id = None
            if selected_equipment_str:
                parts = selected_equipment_str.split("(ID: ")
                if len(parts) > 1:
                    equipment_id = parts[1].strip(")")

            action_taken = action_taken_entry.get()

            if not all([date, time, event_type, description, action_taken]):
                messagebox.showwarning("Eksik Bilgi", "Lütfen tüm zorunlu alanları doldurun.")
                return

            try:
                self.db.execute_query("""
                    INSERT INTO emergency_events (date, time, event_type, description, affected_personnel_ids, affected_equipment_ids, action_taken)
                    VALUES (?, ?, ?, ?, ?, ?, ?)
                """, (date, time, event_type, description, personnel_id, equipment_id, action_taken))
                messagebox.showinfo("Başarılı", "Acil durum kaydı başarıyla eklendi.")
                self.load_emergency_data() # Listeyi güncelle
                add_window.destroy() # Pencereyi kapat
            except Exception as e:
                messagebox.showerror("Hata", f"Acil durum kaydı eklenirken bir hata oluştu: {e}")

        # Kaydet butonu (image_652ae5.png'deki yeşil butona benzer stil)
        save_button = tk.Button(add_window, text="Kaydet", command=save_emergency,
                                bg="#2ecc71", fg="white", font=("Arial", 10, "bold"),
                                padx=15, pady=8, bd=0, relief="flat")
        save_button.pack(pady=20)

    def edit_emergency(self):
        selected_item = self.emergency_tree.focus()
        if not selected_item:
            messagebox.showwarning("Seçim Yok", "Lütfen düzenlemek için bir acil durum kaydı seçin.")
            return
        
        values = self.emergency_tree.item(selected_item, 'values')
        emergency_id = values[0] # ID her zaman ilk sütunda

        # Düzenleme penceresi oluştur
        edit_window = tk.Toplevel(self.root)
        edit_window.title("Acil Durum Kaydını Düzenle")
        edit_window.geometry("400x550") # Tasarıma uygun boyut
        edit_window.transient(self.root)
        edit_window.grab_set()
        edit_window.configure(bg="#ecf0f1") # Arka plan rengini ekledim

        # Etiket ve giriş alanı stilleri için varsayılan font ve genişlikleri kullanıyorum.
        label_font = ("Arial", 10)
        entry_font = ("Arial", 10)
        entry_width = 37 # Tasarıma uygun genişlik
        pady_value = 5 # Boşluk değeri

        tk.Label(edit_window, text="Tarih (YYYY-MM-DD):", font=label_font, bg="#ecf0f1").pack(pady=(10,0))
        date_entry = tk.Entry(edit_window, font=entry_font, width=entry_width)
        date_entry.insert(0, values[1])
        date_entry.pack(pady=pady_value)

        tk.Label(edit_window, text="Saat (HH:MM):", font=label_font, bg="#ecf0f1").pack(pady=(10,0))
        time_entry = tk.Entry(edit_window, font=entry_font, width=entry_width)
        time_entry.insert(0, values[2])
        time_entry.pack(pady=pady_value)

        tk.Label(edit_window, text="Acil Durum Türü:", font=label_font, bg="#ecf0f1").pack(pady=(10,0))
        event_type_entry = tk.Entry(edit_window, font=entry_font, width=entry_width)
        event_type_entry.insert(0, values[3])
        event_type_entry.pack(pady=pady_value)

        tk.Label(edit_window, text="Açıklama:", font=label_font, bg="#ecf0f1").pack(pady=(10,0))
        description_entry = tk.Entry(edit_window, font=entry_font, width=entry_width)
        description_entry.insert(0, values[4])
        description_entry.pack(pady=pady_value)

        # --- Etkilenen Personel Combobox ---
        tk.Label(edit_window, text="Etkilenen Personel:", font=label_font, bg="#ecf0f1").pack(pady=(10,0))
        personnel_data = self.db.fetch_all("SELECT id, name, surname FROM personnel")
        personnel_options = [""] + [f"{p[1]} {p[2]} (ID: {p[0]})" for p in personnel_data]
        personnel_combobox = ttk.Combobox(edit_window, values=personnel_options, state="readonly", width=entry_width - 2)
        personnel_combobox.pack(pady=pady_value)

        # Mevcut personel bilgisini seçili hale getir
        # values[5] -> load_emergency_data'dan gelen "Ad Soyad (ID: X)" veya boş string
        personnel_combobox.set(values[5])

        # --- Etkilenen Ekipman Combobox ---
        tk.Label(edit_window, text="Etkilenen Ekipman:", font=label_font, bg="#ecf0f1").pack(pady=(10,0))
        equipment_data = self.db.fetch_all("SELECT id, name FROM equipment")
        equipment_options = [""] + [f"{e[1]} (ID: {e[0]})" for e in equipment_data]
        equipment_combobox = ttk.Combobox(edit_window, values=equipment_options, state="readonly", width=entry_width - 2)
        equipment_combobox.pack(pady=pady_value)
        
        # Mevcut ekipman bilgisini seçili hale getir
        # values[6] -> load_emergency_data'dan gelen "Ekipman Adı (ID: X)" veya boş string
        equipment_combobox.set(values[6])

        tk.Label(edit_window, text="Alınan Aksiyon:", font=label_font, bg="#ecf0f1").pack(pady=(10,0))
        action_taken_entry = tk.Entry(edit_window, font=entry_font, width=entry_width)
        action_taken_entry.insert(0, values[7])
        action_taken_entry.pack(pady=pady_value)

        def save_edited_emergency():
            date = date_entry.get()
            time = time_entry.get()
            event_type = event_type_entry.get()
            description = description_entry.get()

            # Combobox'lardan ID'leri al
            selected_personnel_str = personnel_combobox.get()
            personnel_id = None
            if selected_personnel_str:
                parts = selected_personnel_str.split("(ID: ")
                if len(parts) > 1:
                    personnel_id = parts[1].strip(")")

            selected_equipment_str = equipment_combobox.get()
            equipment_id = None
            if selected_equipment_str:
                parts = selected_equipment_str.split("(ID: ")
                if len(parts) > 1:
                    equipment_id = parts[1].strip(")")
            
            action_taken = action_taken_entry.get()

            if not all([date, time, event_type, description, action_taken]):
                messagebox.showwarning("Eksik Bilgi", "Lütfen tüm zorunlu alanları doldurun.")
                return

            try:
                # UPDATE sorgusunu Combobox'tan gelen ID'lerle güncelledim
                self.db.execute_query("""
                    UPDATE emergency_events
                    SET date=?, time=?, event_type=?, description=?, affected_personnel_ids=?, affected_equipment_ids=?, action_taken=?
                    WHERE id=?
                """, (date, time, event_type, description, personnel_id, equipment_id, action_taken, emergency_id))
                messagebox.showinfo("Başarılı", "Acil durum kaydı başarıyla güncellendi.")
                self.load_emergency_data() # Listeyi güncelle
                edit_window.destroy()
            except Exception as e:
                messagebox.showerror("Hata", f"Acil durum kaydı güncellenirken bir hata oluştu: {e}")

        # Kaydet butonu (image_652ae5.png'deki yeşil butona benzer stil)
        save_button = tk.Button(edit_window, text="Kaydet", command=save_edited_emergency,
                                bg="#2ecc71", fg="white", font=("Arial", 10, "bold"),
                                padx=15, pady=8, bd=0, relief="flat")
        save_button.pack(pady=20)

    def delete_emergency(self):
        selected_item = self.emergency_tree.focus()
        if not selected_item:
            messagebox.showwarning("Seçim Yok", "Lütfen silmek için bir acil durum kaydı seçin.")
            return

        confirm = messagebox.askyesno("Silme Onayı", "Seçili acil durum kaydını silmek istediğinizden emin misiniz?")
        if confirm:
            values = self.emergency_tree.item(selected_item, 'values')
            emergency_id = values[0]
            try:
                self.db.execute_query("DELETE FROM emergency_events WHERE id=?", (emergency_id,))
                messagebox.showinfo("Başarılı", "Acil durum kaydı başarıyla silindi.")
                self.load_emergency_data() # Listeyi güncelle
            except Exception as e:
                messagebox.showerror("Hata", f"Acil durum kaydı silinirken bir hata oluştu: {e}")
    # --- Raporlar Modülü ---
    def show_reports_module(self):
        self.clear_content_frame() # Önceki içeriği temizle
        tk.Label(self.content_frame, text="Raporlar Modülü", font=("Arial", 16, "bold"), bg="#ecf0f1").pack(pady=20)
        
        # Rapor seçenekleri çerçevesi
        control_frame = ttk.LabelFrame(self.content_frame, text="Rapor Seçenekleri")
        control_frame.pack(padx=10, pady=10, fill="x", expand=False)

        # Rapor Tipi Seçimi
        ttk.Label(control_frame, text="Rapor Tipi:").grid(row=0, column=0, padx=5, pady=5, sticky="w")
        
        self.rapor_tipleri = [
            "Tüm Personel Listesi",
            "Tüm Ekipman Listesi",
            "Tüm Görevler Listesi",
            "Tüm Acil Durumlar",
            "Personel Görev Raporu",
            "Ekipman Atanmış Personel Raporu"
            # Buraya daha fazla rapor tipi eklenebilir
        ]
        self.rapor_secimi_cb = ttk.Combobox(control_frame, values=self.rapor_tipleri, state="readonly")
        self.rapor_secimi_cb.grid(row=0, column=1, padx=5, pady=5, sticky="ew")
        self.rapor_secimi_cb.set(self.rapor_tipleri[0]) # Varsayılan olarak ilk raporu seç
        self.rapor_secimi_cb.bind("<<ComboboxSelected>>", self.on_rapor_type_select) # Seçim değiştiğinde tetiklenecek

        # Rapor Oluştur Butonu
        ttk.Button(control_frame, text="Rapor Oluştur", command=self.generate_report).grid(row=0, column=2, padx=5, pady=5)
        
        # Rapor Çıktısı Çerçevesi
        output_frame = ttk.LabelFrame(self.content_frame, text="Rapor Çıktısı")
        output_frame.pack(padx=10, pady=5, fill="both", expand=True)

        # Rapor Treeview'ı (çıktı alanı)
        self.report_tree = ttk.Treeview(output_frame, show="headings")
        self.report_tree.pack(side="left", fill="both", expand=True)

        # Scrollbar ekleme
        report_scrollbar_y = ttk.Scrollbar(output_frame, orient="vertical", command=self.report_tree.yview)
        report_scrollbar_y.pack(side="right", fill="y")
        self.report_tree.configure(yscrollcommand=report_scrollbar_y.set)

        report_scrollbar_x = ttk.Scrollbar(output_frame, orient="horizontal", command=self.report_tree.xview)
        report_scrollbar_x.pack(side="bottom", fill="x")
        self.report_tree.configure(xscrollcommand=report_scrollbar_x.set)

        # Dışa Aktar Butonları Çerçevesi
        export_frame = ttk.LabelFrame(self.content_frame, text="Raporu Dışa Aktar")
        export_frame.pack(padx=10, pady=5, fill="x", expand=False)

        ttk.Button(export_frame, text="Excel'e Aktar", command=lambda: self.export_report("excel")).grid(row=0, column=0, padx=5, pady=5)
        ttk.Button(export_frame, text="PDF'e Aktar", command=lambda: self.export_report("pdf")).grid(row=0, column=1, padx=5, pady=5)
        ttk.Button(export_frame, text="Metin Dosyasına Aktar", command=lambda: self.export_report("txt")).grid(row=0, column=2, padx=5, pady=5)
        
        # Seçilen rapor tipine göre Treeview sütunlarını güncellemek için başlangıçta çağır
        self.on_rapor_type_select(None) # None ile çağırabiliriz çünkü event objesine ihtiyacımız yok

    def on_rapor_type_select(self, event):
        # Treeview'ın sütunlarını seçilen rapor tipine göre ayarlar
        selected_report = self.rapor_secimi_cb.get()
        
        # Mevcut sütunları temizle
        self.report_tree.delete(*self.report_tree.get_children())
        self.report_tree["columns"] = ()
        self.report_tree["displaycolumns"] = ()

        if selected_report == "Tüm Personel Listesi":
            columns = ("ID", "Ad", "Rol", "Vardiya")
            widths = (50, 150, 100, 100)
        elif selected_report == "Tüm Ekipman Listesi":
            columns = ("ID", "Ekipman Adı", "Atanan Personel", "Durum")
            widths = (50, 150, 150, 100)
        elif selected_report == "Tüm Görevler Listesi":
            columns = ("ID", "Görev Adı", "Atanan Kişi", "Son Teslim Tarihi", "Durum")
            widths = (50, 200, 150, 100, 100)
        elif selected_report == "Tüm Acil Durumlar":
            columns = ("ID", "Tarih", "Saat", "Olay Tipi", "Açıklama", "Etkilenen Personel", "Etkilenen Ekipman", "Alınan Önlem")
            widths = (50, 100, 80, 120, 200, 200, 200, 200)
        elif selected_report == "Personel Görev Raporu":
            columns = ("Personel Adı", "Görev Adı", "Son Teslim Tarihi", "Durum")
            widths = (150, 200, 100, 100)
        elif selected_report == "Ekipman Atanmış Personel Raporu":
            columns = ("Ekipman Adı", "Atanan Personel", "Ekipman Durumu")
            widths = (150, 150, 100)
        else: # Varsayılan veya bilinmeyen rapor tipi
            columns = ("Bilgi")
            widths = (300)

        self.report_tree["columns"] = columns
        for col, width in zip(columns, widths):
            self.report_tree.heading(col, text=col)
            self.report_tree.column(col, width=width, anchor="w")

    def generate_report(self):
        # Mevcut verileri temizle
        for item in self.report_tree.get_children():
            self.report_tree.delete(item)

        selected_report = self.rapor_secimi_cb.get()
        data = []

        if selected_report == "Tüm Personel Listesi":
            data = self.db.fetch_all("SELECT id, name, role, shift FROM personnel")
        elif selected_report == "Tüm Ekipman Listesi":
            # Ekipman için atanan personelin adını da çekmek için JOIN
            data = self.db.fetch_all("""
                SELECT
                    e.id,
                    e.name,
                    p.name AS assigned_personnel_name,
                    e.status
                FROM
                    equipment e
                LEFT JOIN
                    personnel p ON e.assigned_to = p.id
            """)
        elif selected_report == "Tüm Görevler Listesi":
            # Görevler için atanan kişinin adını da çekmek için JOIN
            data = self.db.fetch_all("""
                SELECT
                    t.id,
                    t.task_name,
                    p.name AS assigned_person_name,
                    t.deadline,
                    t.status
                FROM
                    tasks t
                LEFT JOIN
                    personnel p ON t.assigned_person = p.id
            """)
        elif selected_report == "Tüm Acil Durumlar":
            # Acil Durumlar için personel ve ekipman adlarını Python'da işlemek üzere ID'leri çek
            raw_emergency_data = self.db.fetch_all("""
                SELECT
                    ee.id,
                    ee.date,
                    ee.time,
                    ee.event_type,
                    ee.description,
                    ee.affected_personnel_ids,
                    ee.affected_equipment_ids,
                    ee.action_taken
                FROM
                    emergency_events ee
                ORDER BY
                    ee.date DESC, ee.time DESC
            """)
            
            # Python'da işleme (önceki düzeltmelerimizdeki gibi)
            for event in raw_emergency_data:
                event_id, date, time, event_type, description, personnel_ids_str, equipment_ids_str, action_taken = event

                affected_personnel_display = ""
                if personnel_ids_str: # None veya boş string kontrolü
                    personnel_ids = []
                    for p_id_str in personnel_ids_str.split(','):
                        try:
                            personnel_ids.append(int(p_id_str.strip()))
                        except ValueError:
                            continue
                    
                    personel_names = []
                    for p_id in personnel_ids:
                        personnel_info = self.db.get_personnel_by_id(p_id)
                        if personnel_info:
                            personel_names.append(f"{personnel_info[1]} (ID: {p_id})") # personnel_info[1] is 'name'
                    affected_personnel_display = ", ".join(personel_names)
                
                affected_equipment_display = ""
                if equipment_ids_str: # None veya boş string kontrolü
                    equipment_ids = []
                    for e_id_str in equipment_ids_str.split(','):
                        try:
                            equipment_ids.append(int(e_id_str.strip()))
                        except ValueError:
                            continue

                    equipment_names = []
                    for e_id in equipment_ids:
                        equipment_info = self.db.get_equipment_by_id(e_id)
                        if equipment_info:
                            equipment_names.append(f"{equipment_info[1]} (ID: {e_id})") # equipment_info[1] is 'name'
                    affected_equipment_display = ", ".join(equipment_names)
                
                data.append((event_id, date, time, event_type, description, affected_personnel_display, affected_equipment_display, action_taken))

        elif selected_report == "Personel Görev Raporu":
            data = self.db.fetch_all("""
                SELECT
                    p.name AS personnel_name,
                    t.task_name,
                    t.deadline,
                    t.status
                FROM
                    tasks t
                LEFT JOIN
                    personnel p ON t.assigned_person = p.id
                ORDER BY
                    p.name, t.deadline
            """)
        elif selected_report == "Ekipman Atanmış Personel Raporu":
            data = self.db.fetch_all("""
                SELECT
                    e.name AS equipment_name,
                    p.name AS assigned_personnel_name,
                    e.status
                FROM
                    equipment e
                LEFT JOIN
                    personnel p ON e.assigned_to = p.id
                ORDER BY
                    e.name
            """)
        # Treeview'a verileri ekle
        for row in data:
            self.report_tree.insert("", "end", values=row)

        if not data:
            self.report_tree.insert("", "end", values=("Gösterilecek veri bulunamadı.",))


    def export_report(self, format_type):
        selected_report = self.rapor_secimi_cb.get()
        # Treeview'daki tüm veriyi al
        headers = [self.report_tree.heading(col)["text"] for col in self.report_tree["columns"]]
        data_to_export = [self.report_tree.item(item)["values"] for item in self.report_tree.get_children()]

        if not data_to_export:
            messagebox.showinfo("Bilgi", "Dışa aktarılacak veri bulunamadı.")
            return

        file_name = f"{selected_report.replace(' ', '_')}_Rapor_{datetime.now().strftime('%Y%m%d_%H%M%S')}"
        
        if format_type == "excel":
            try:
                # openpyxl kütüphanesini yüklemeniz gerekebilir: pip install openpyxl
                from openpyxl import Workbook # BURAYA DİKKAT: Fonksiyon içinde import!
                wb = Workbook()
                ws = wb.active
                ws.title = selected_report

                ws.append(headers) # Başlıkları ekle
                for row in data_to_export:
                    ws.append(row)
                
                filepath = filedialog.asksaveasfilename(defaultextension=".xlsx",
                                                        filetypes=[("Excel files", "*.xlsx")],
                                                        initialfile=file_name)
                if filepath:
                    wb.save(filepath)
                    messagebox.showinfo("Başarılı", f"Rapor Excel olarak kaydedildi:\n{filepath}")
            except ImportError:
                messagebox.showerror("Hata", "openpyxl kütüphanesi yüklü değil. 'pip install openpyxl' komutunu çalıştırın.")
            except Exception as e:
                messagebox.showerror("Hata", f"Excel'e aktarılırken bir hata oluştu: {e}")

        # gui.py dosyasında, export_report fonksiyonunun içinde, format_type == "pdf" bloğu
    def export_report(self, format_type):
        selected_report = self.rapor_secimi_cb.get()
        headers = [self.report_tree.heading(col)["text"] for col in self.report_tree["columns"]]
        data_to_export = [self.report_tree.item(item)["values"] for item in self.report_tree.get_children()]

        if not data_to_export:
            messagebox.showinfo("Bilgi", "Dışa aktarılacak veri bulunamadı.")
            return

        file_name = f"{selected_report.replace(' ', '_')}_Rapor_{datetime.now().strftime('%Y%m%d_%H%M%S')}"
        
        if format_type == "excel":
            try:
                from openpyxl import Workbook
                wb = Workbook()
                ws = wb.active
                ws.title = selected_report

                ws.append(headers)
                for row in data_to_export:
                    ws.append(row)
                
                filepath = filedialog.asksaveasfilename(defaultextension=".xlsx",
                                                        filetypes=[("Excel files", "*.xlsx")],
                                                        initialfile=file_name)
                if filepath:
                    wb.save(filepath)
                    messagebox.showinfo("Başarılı", f"Rapor Excel olarak kaydedildi:\n{filepath}")
            except ImportError:
                messagebox.showerror("Hata", "openpyxl kütüphanesi yüklü değil. 'pip install openpyxl' komutunu çalıştırın.")
            except Exception as e:
                messagebox.showerror("Hata", f"Excel'e aktarılırken bir hata oluştu: {e}")

        elif format_type == "pdf": # Bu blok kendi içinde doğru girintiye sahip
            try:
                from fpdf import FPDF 
                
                pdf = FPDF()
                pdf.add_page()
                
                font_path = "DejaVuSansCondensed.ttf" # Veya "fonts/DejaVuSansCondensed.ttf"

                pdf.add_font("DejaVuSans", "", font_path, uni=True)
                pdf.set_font("DejaVuSans", size=10)

                pdf.cell(200, 10, txt=selected_report, ln=True, align="C")
                pdf.ln(5)

                col_widths = []
                for h in headers:
                    col_widths.append(pdf.get_string_width(h) + 10) 
                
                total_width = sum(col_widths)
                page_width = pdf.w - 2*pdf.l_margin
                scale_factor = page_width / total_width
                
                adjusted_col_widths = [w * scale_factor for w in col_widths]

                for i, header in enumerate(headers):
                    pdf.cell(adjusted_col_widths[i], 10, header, 1, 0, 'C')
                pdf.ln()

                for row_data in data_to_export:
                    for i, cell_data in enumerate(row_data):
                        cell_text = str(cell_data) 
                        pdf.cell(adjusted_col_widths[i], 10, cell_text, 1, 0, 'L')
                    pdf.ln()

                filepath = filedialog.asksaveasfilename(defaultextension=".pdf",
                                                        filetypes=[("PDF files", "*.pdf")],
                                                        initialfile=file_name)
                if filepath:
                    pdf.output(filepath)
                    messagebox.showinfo("Başarılı", f"Rapor PDF olarak kaydedildi:\n{filepath}")

            except ImportError:
                messagebox.showerror("Hata", "fpdf kütüphanesi yüklü değil. 'pip install fpdf' komutunu çalıştırın.")
            except FileNotFoundError:
                messagebox.showerror("Hata", f"Font dosyası bulunamadı: {font_path}\nLütfen font dosyasının doğru yolda olduğundan emin olun.")
            except Exception as e:
                messagebox.showerror("Hata", f"PDF'e aktarılırken bir hata oluştu: {e}")

        elif format_type == "txt": # Bu blok da önceki if/elif ile aynı hizadan başlamalı
            try:
                filepath = filedialog.asksaveasfilename(defaultextension=".txt",
                                                        filetypes=[("Text files", "*.txt")],
                                                        initialfile=file_name)
                if filepath:
                    with open(filepath, "w", encoding="utf-8") as f:
                        f.write("\t".join(headers) + "\n") # Başlıkları tab ile ayır
                        for row in data_to_export:
                            f.write("\t".join(map(str, row)) + "\n") # Verileri tab ile ayır
                    messagebox.showinfo("Başarılı", f"Rapor metin dosyası olarak kaydedildi:\n{filepath}")
            except Exception as e:
                messagebox.showerror("Hata", f"Metin dosyasına aktarılırken bir hata oluştu: {e}")